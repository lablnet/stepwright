# handlers/data_flow_handlers.py
# Advanced data flow handlers for StepWright (File I/O and Custom Callbacks)
# Author: Muhammad Umer Farooq <umer@lablnet.com>

import csv
import json
import pathlib
from typing import Any, Dict

from playwright.async_api import Page
from ..step_types import BaseStep
from ..helpers import replace_data_placeholders, maybe_await


async def _handle_read_data(
    page: Page,
    step: BaseStep,
    collector: Dict[str, Any],
) -> None:
    """Handle readData action (JSON, CSV, Excel, Text)"""
    path = replace_data_placeholders(step.value or "", collector)
    if not path:
        raise ValueError(
            f"readData requires a file path in 'value' for step '{step.id}'"
        )

    file_path = pathlib.Path(path)
    if not file_path.exists() and step.data_type != "custom":
        if step.continueOnEmpty:
            print(f"   ⚠️  File not found: {path} - skipping")
            return
        raise FileNotFoundError(f"File not found: {path}")

    data: Any = None
    fmt = step.data_type or "text"

    try:
        if fmt == "json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        elif fmt == "csv":
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                data = list(reader)
        elif fmt == "excel":
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                data = []
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data.append(dict(zip(headers, row)))
            except ImportError:
                raise ImportError(
                    "openpyxl is required for Excel support. Install with 'pip install openpyxl'"
                )
        elif fmt == "custom" and step.callback:
            data = await maybe_await(step.callback(path, step))
        else:  # Default to text/lines
            with open(file_path, "r", encoding="utf-8") as f:
                data = [line.strip() for line in f if line.strip()]

        key = step.key or step.id
        collector[key] = data
        print(
            f"   📂 Read {len(data) if isinstance(data, list) else 1} items from {path}"
        )

    except Exception as e:
        if step.skipOnError:
            print(f"   ⚠️  readData failed for {path} (skipping): {e}")
        else:
            raise


async def _handle_write_data(
    page: Page,
    step: BaseStep,
    collector: Dict[str, Any],
) -> None:
    """Handle writeData action (JSON, CSV, Excel, Text)"""
    path = replace_data_placeholders(step.value or "", collector)
    if not path:
        raise ValueError(
            f"writeData requires a file path in 'value' for step '{step.id}'"
        )

    data_to_write = collector.get(step.key) if step.key else collector
    if data_to_write is None:
        print(f"   ⚠️  No data found for key '{step.key}' - skipping write")
        return

    file_path = pathlib.Path(path)
    # Ensure directory exists
    file_path.parent.mkdir(parents=True, exist_ok=True)

    fmt = step.data_type or "text"
    append = step.continueOnEmpty is not False  # Reuse this flag or similar?

    try:
        if fmt == "json":
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data_to_write, f, indent=2)
        elif fmt == "csv":
            if not isinstance(data_to_write, list):
                data_to_write = [data_to_write]
            if not data_to_write:
                return

            headers = data_to_write[0].keys()
            write_header = not file_path.exists()
            with open(file_path, "a", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                if write_header:
                    writer.writeheader()
                writer.writerows(data_to_write)
        elif fmt == "excel":
            try:
                import openpyxl
                from openpyxl import Workbook

                if not isinstance(data_to_write, list):
                    data_to_write = [data_to_write]

                if file_path.exists():
                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active
                else:
                    wb = Workbook()
                    sheet = wb.active
                    if data_to_write:
                        sheet.append(list(data_to_write[0].keys()))

                for item in data_to_write:
                    sheet.append(list(item.values()))
                wb.save(file_path)
            except ImportError:
                raise ImportError(
                    "openpyxl is required for Excel support. Install with 'pip install openpyxl'"
                )
        elif fmt == "custom" and step.callback:
            await maybe_await(step.callback(path, data_to_write, step))
        else:  # Text
            with open(file_path, "a" if append else "w", encoding="utf-8") as f:
                if isinstance(data_to_write, list):
                    for item in data_to_write:
                        f.write(f"{item}\n")
                else:
                    f.write(f"{data_to_write}\n")

        print(f"   💾 Saved data to {path}")

    except Exception as e:
        if step.skipOnError:
            print(f"   ⚠️  writeData failed for {path} (skipping): {e}")
        else:
            raise


async def _handle_custom_callback(
    page: Page,
    step: BaseStep,
    collector: Dict[str, Any],
) -> None:
    """Handle custom action using the provided callback"""
    if not step.callback:
        raise ValueError(f"Action 'custom' requires a 'callback' in step '{step.id}'")

    print(f"   ⚡ Executing custom callback for '{step.id}'")
    result = await maybe_await(step.callback(page, collector, step))

    if step.key and result is not None:
        collector[step.key] = result
